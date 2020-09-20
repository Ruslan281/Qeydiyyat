import tkinter as tk
import tkinter.ttk as ttk

from openpyxl import *
from tkinter.messagebox import showinfo

win=tk.Tk()
win.title("Qeydiyyat")



def save():
    ad=entry.get()
    familya=entry1.get()
    yas=entry2.get()

    wb=Workbook()
    ws=wb.active

    ws["A1"] = "Adiniz"
    ws["B1"] = "Familyaniz"
    ws["C1"] = "Yasiniz"
    ws["A3"] = ad
    ws["B3"] = familya
    ws["C3"] = yas

    wb.save(r"C:\Users\KTB\Desktop\\qeydiyyatim.xlsx")
    showinfo("Save","Yaddasda saxlanildi")


def clear():

    entry.delete(0,tk.END)
    entry1.delete(0,tk.END)
    entry2.delete(0,tk.END)


label=tk.Label(win,text="Adiniz : ")
label.grid(row=0,column=0,padx=8,pady=8)

entry=tk.Entry(win)
entry.grid(row=0,column=1,padx=8,pady=8)

label1=tk.Label(win,text="Familiya  : ")
label1.grid(row=1,column=0,padx=8,pady=8)

entry1=tk.Entry(win)
entry1.grid(row=1,column=1,padx=8,pady=8)

label2=tk.Label(win,text="Yasiniz  : ")
label2.grid(row=2,column=0,padx=8,pady=8)

entry2=tk.Entry(win)
entry2.grid(row=2,column=1,padx=8,pady=8)

button=tk.Button(win,text="Qeydiyyat",command=save)
button.grid(row=3,column=0,padx=8,pady=8)

button2=tk.Button(win,text="Temizle",command=clear)
button2.grid(row=3,column=1,padx=8,pady=8)

win.mainloop()



















